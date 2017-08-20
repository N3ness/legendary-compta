from openpyxl import Workbook
import time
class Operations:

    # def __copyToWorksheet(self):


    def createJournalReport(self,journalList,path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Journal"
        ws['A1']= 'Données exportées le ' + str(time.strftime("%d/%m/%Y"))

        rowterator = 5
        for row in journalList:
            columnterator=2
            for cell in row:
                ws.cell(row=rowterator, column=columnterator,value=cell)
                columnterator +=1
            rowterator+=1
            print(cell)

        wb.save(path)

